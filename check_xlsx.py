import openpyxl
wb = openpyxl.load_workbook('cricviz_all_pages.xlsx')
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'\nSheet: {name}, Dimensions: {ws.dimensions}, Rows: {ws.max_row}, Cols: {ws.max_column}')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
        print(f'  Row {i}: {row[:10]}')  # first 10 cols
